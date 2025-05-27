import os
import io
import glob
import zipfile
import tempfile
from datetime import datetime
from flask import Blueprint, render_template, request, redirect, url_for, flash, session, send_file, current_app
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from openpyxl import load_workbook
from openpyxl.cell import MergedCell
from bson.objectid import ObjectId

from app import allowed_file

# Inicializar Blueprint
main = Blueprint('main', __name__)

# Set para hojas actualizadas
hojas_actualizadas = set()

# Función para obtener conexión Mongo y sus colecciones
def get_db():
    return current_app.config['MONGO_DB']

def get_collections():
    db = get_db()
    return {
        'registros': db['registros'],
        'usuarios': db['usuarios'],
        'alertas': db['alertas']
    }

# =================== RUTAS ===================

@main.route('/')
def home():
    return render_template('index.html')  # donde están los 3 botones


@main.route('/dashboard')
def dashboard():
    if 'usuario' not in session:
        return redirect(url_for('main.login'))
    return render_template('dashboard.html', usuario=session['usuario'])




@main.route('/login', methods=['GET', 'POST'])
def login():
    col = get_collections()['usuarios']

    if request.method == 'POST':
        usuario = request.form['usuario']
        password = request.form['password']
        user = col.find_one({"usuario": usuario})

        if user:
            if not user.get('activo', True):
                mensaje = "❌ Usuario desactivado. Contacte al administrador."
                return render_template('login.html', mensaje=mensaje)

            if check_password_hash(user['password'], password):
                session['usuario'] = usuario
                session['rol'] = user.get('rol', 'usuario')
                return redirect(url_for('main.dashboard'))  # 👈 nuevo destino
            else:
                mensaje = "❌ Usuario o contraseña incorrectos."
                return render_template('login.html', mensaje=mensaje)

        else:
            mensaje = "❌ Usuario o contraseña incorrectos."
            return render_template('login.html', mensaje=mensaje)

    return render_template('login.html')





@main.route('/register', methods=['GET', 'POST'])
def register():
    col = get_collections()['usuarios']
    if request.method == 'POST':
        usuario = request.form['usuario']
        email = request.form['email']
        password = request.form['password']

        if col.find_one({"usuario": usuario}):
            return render_template('register.html', mensaje="❌ El nombre de usuario ya está registrado.")
        if col.find_one({"email": email}):
            return render_template('register.html', mensaje="❌ El correo ya está registrado.")

        col.insert_one({
            "usuario": usuario,
            "email": email,
            "password": generate_password_hash(password),
            "rol": "usuario",
            "activo": True,
            "creado": datetime.now()
        })

        return render_template('register.html', mensaje="Usuario registrado exitosamente. Ahora puede iniciar sesión.")

    return render_template('register.html')



@main.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    col = get_collections()['usuarios']
    
    if request.method == 'POST':
        usuario = request.form['usuario']
        user = col.find_one({"usuario": usuario})

        if not user:
            mensaje = "❌ Usuario no encontrado."
            return render_template('forgot_password.html', mensaje=mensaje)

        # Aquí podrías enviar un correo, pero por ahora redirigimos a nueva página:
        return redirect(url_for('main.restablecer_contrasena', usuario=usuario))

    return render_template('forgot_password.html')


@main.route('/reset-password/<usuario>', methods=['GET', 'POST'])
def restablecer_contrasena(usuario):
    col = get_collections()['usuarios']

    if request.method == 'POST':
        nueva_password = request.form['password']
        col.update_one(
            {"usuario": usuario},
            {"$set": {"password": generate_password_hash(nueva_password)}}
        )
        flash("✅ Contraseña actualizada correctamente.")
        return redirect(url_for('main.login'))

    return render_template('reset_password.html', usuario=usuario)





@main.route('/alertas', methods=['GET', 'POST'])
def alertas():
    col = get_collections()['alertas']
    if request.method == 'POST':
        alerta = {
            "usuario": session.get('usuario'),
            "fecha": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            "serie_computador": request.form['serie'],
            "tipo_ubicacion": request.form['tipo_ubicacion'],
            "ubicacion": request.form['ubicacion'],
            "problema": request.form['problema']
        }
        col.insert_one(alerta)
        return render_template('alertas.html', mensaje="✅ Alerta registrada con éxito.")

    return render_template('alertas.html')

@main.route('/hojas-de-vida')
def hojas_de_vida():
    archivos = sorted(hojas_actualizadas)
    return render_template('hojas_de_vida.html', archivos=archivos)

from bson.objectid import ObjectId



@main.route('/usuarios')
def usuarios():
    if 'usuario' not in session or session.get('rol') != 'admin':
        return redirect(url_for('main.login'))

    col = get_collections()['usuarios']
    usuarios = list(col.find())
    return render_template('usuarios.html', usuarios=usuarios)




@main.route('/cambiar-estado-usuario', methods=['POST'])
def cambiar_estado_usuario():
    if 'usuario' not in session or session.get('rol') != 'admin':
        return redirect(url_for('main.login'))

    col = get_collections()['usuarios']
    usuario_id = request.form.get('id')
    usuario = col.find_one({"_id": ObjectId(usuario_id)})

    if usuario:
        nuevo_estado = not usuario.get('activo', True)
        col.update_one({"_id": ObjectId(usuario_id)}, {"$set": {"activo": nuevo_estado}})

    return redirect(url_for('main.usuarios'))



@main.route('/limpiar-historial', methods=['POST'])
def limpiar_historial():
    col = get_collections()['registros']
    col.delete_many({})
    hojas_actualizadas.clear()
    return redirect(url_for('main.hojas_de_vida'))

@main.route('/plan-general')
def plan_general():
    return render_template('plan_general.html')

@main.route('/mantenimiento', methods=['GET', 'POST'])
def mantenimiento():
    col = get_collections()['registros']
    if 'usuario' not in session:
        return redirect(url_for('main.login'))

    if request.method == 'POST':
        equipo = request.form['equipo']
        fecha = request.form['fecha']
        descripcion = request.form['descripcion']

        col.insert_one({"equipo": equipo, "fecha": fecha, "descripcion": descripcion})
        actualizar_hoja_vida(equipo, fecha, descripcion)
        hojas_actualizadas.add(f"{equipo}.xlsx")

        return redirect(url_for('main.mantenimiento', mensaje='ok'))

    mensaje = request.args.get('mensaje')
    registros = list(col.find().sort("_id", -1))
    return render_template('mantenimiento.html', registros=registros, mensaje=mensaje)

@main.route('/mantenimientos-mensuales')
def mantenimientos_mensuales():
    col = get_collections()['registros']
    registros = list(col.find())
    registros_por_mes = {}

    for reg in registros:
        try:
            fecha = datetime.strptime(reg['fecha'], '%Y-%m-%d')
        except Exception as e:
            continue

        mes = fecha.strftime('%Y-%m')
        registros_por_mes.setdefault(mes, []).append(reg)

    return render_template('mantenimientos_mensuales.html', registros_por_mes=registros_por_mes)

@main.route('/reportes')
def reportes():
    if 'usuario' not in session:
        return redirect(url_for('main.login'))

    db = get_db()  # ← CORRECTO
    usuario_actual = session['usuario']
    alertas_activas = list(db.alertas.find({"usuario": usuario_actual}))
    alertas_resueltas = list(db.alertas_resueltas.find({"usuario": usuario_actual}))

    alertas_usuario = sorted(alertas_activas + alertas_resueltas, key=lambda x: x['fecha'], reverse=True)

    return render_template('reportes.html', alertas=alertas_usuario)

@main.route('/limpiar-reportes', methods=['POST'])
def limpiar_reportes():
    if 'usuario' not in session:
        return redirect(url_for('main.login'))

    db = get_db()
    usuario = session['usuario']

    db.alertas.delete_many({'usuario': usuario})
    db.alertas_resueltas.delete_many({'usuario': usuario})

    flash("🧹 Todos tus reportes fueron eliminados correctamente.")
    return redirect(url_for('main.reportes'))



@main.route('/gestionar-alertas')
def gestionar_alertas():
    if 'usuario' not in session or session.get('rol') != 'admin':
        return redirect(url_for('main.login'))

    estado = request.args.get('estado')
    filtro = {"estado": estado} if estado and estado != 'Todas' else {}
    col = get_collections()['alertas']
    alertas = list(col.find(filtro).sort("fecha", -1))
    return render_template('gestionar_alertas.html', alertas=alertas, estado_actual=estado or 'Todas')

@main.route('/actualizar-alerta/<id>', methods=['POST'])
def actualizar_alerta(id):
    if 'usuario' not in session or session.get('rol') != 'admin':
        return redirect(url_for('main.login'))

    col_alertas = get_collections()['alertas']
    col_resueltas = get_db()['alertas_resueltas']

    estado = request.form.get('estado')
    hora_atencion = request.form.get('hora_atencion')
    comentario_admin = request.form.get('comentario_admin')

    alerta = col_alertas.find_one({'_id': ObjectId(id)})

    if alerta:
        alerta['estado'] = estado
        alerta['hora_atencion'] = hora_atencion
        alerta['comentario_admin'] = comentario_admin
    if estado == "Solucionado":
       alerta['_id'] = ObjectId()  # Genera nuevo ID
       col_resueltas.insert_one(alerta)
       col_alertas.delete_one({'_id': ObjectId(id)})

    else:
            col_alertas.update_one(
                {'_id': ObjectId(id)},
                {'$set': {
                    'estado': estado,
                    'hora_atencion': hora_atencion,
                    'comentario_admin': comentario_admin
                }}
            )

    return redirect(url_for('main.gestionar_alertas'))

@main.route('/historial-soportes')
def historial_soportes():
    if 'usuario' not in session or session.get('rol') != 'admin':
        return redirect(url_for('main.login'))

    col = get_db()['alertas_resueltas']
    fecha_filtrada = request.args.get('fecha')  # formato: "2025-05-27"

    filtro = {}
    if fecha_filtrada:
        filtro['fecha'] = {"$regex": fecha_filtrada}

    historial = list(col.find(filtro).sort("fecha", -1))
    return render_template('alertas_resueltas.html', historial=historial, fecha=fecha_filtrada or "")



@main.route('/solucionar-todo', methods=['POST'])
def solucionar_todo():
    if 'usuario' not in session or session.get('rol') != 'admin':
        return redirect(url_for('main.login'))

    estado_filtro = request.form.get('estado_actual')
    filtro = {"estado": estado_filtro} if estado_filtro and estado_filtro != 'Todas' else {"estado": {'$in': ['Pendiente', None]}}
    col = get_collections()['alertas']
    col.update_many(filtro, {
        "$set": {
            "estado": "Solucionado",
            "hora_atencion": datetime.now().strftime('%H:%M'),
            "comentario_admin": "Solucionado por administrador"
        }
    })
    return redirect(url_for('main.gestionar_alertas', estado=estado_filtro))

@main.route('/subir-soporte', methods=['POST'])
def subir_soporte():
    if 'usuario' not in session:
        return redirect(url_for('main.login'))

    if 'archivo' not in request.files:
        flash("❌ No se envió ningún archivo.")
        return redirect(url_for('main.reportes'))

    archivo = request.files['archivo']
    equipo = request.form.get('equipo')

    if archivo.filename == '':
        flash("❌ No se seleccionó ningún archivo.")
        return redirect(url_for('main.reportes'))

    if archivo and allowed_file(archivo.filename):
        nombre = f"{equipo}_{secure_filename(archivo.filename)}"
        ruta = os.path.join(current_app.config['UPLOAD_FOLDER'], nombre)
        archivo.save(ruta)
        flash("✅ Documento subido correctamente.")
    else:
        flash("❌ Tipo de archivo no permitido. Solo se acepta .docx.")

    return redirect(url_for('main.reportes'))

@main.route('/documentos-subidos')
def documentos_subidos():
    if 'usuario' not in session or session.get('rol') != 'admin':
        return redirect(url_for('main.login'))

    carpeta = current_app.config['UPLOAD_FOLDER']
    archivos = sorted(os.listdir(carpeta), reverse=True)
    return render_template('documentos_subidos.html', archivos=archivos)

@main.route('/descargar-todos-los-soportes')
def descargar_todos_los_soportes():
    if 'usuario' not in session or session.get('rol') != 'admin':
        return redirect(url_for('main.login'))

    carpeta = current_app.config['UPLOAD_FOLDER']
    archivos = os.listdir(carpeta)

    if not archivos:
        flash("No hay archivos para descargar.")
        return redirect(url_for('main.documentos_subidos'))

    zip_temp = tempfile.NamedTemporaryFile(delete=False, suffix='.zip')
    with zipfile.ZipFile(zip_temp.name, 'w') as zipf:
        for archivo in archivos:
            ruta = os.path.join(carpeta, archivo)
            zipf.write(ruta, arcname=archivo)

    return send_file(zip_temp.name, as_attachment=True, download_name="soportes_tecnicos.zip")

@main.route('/eliminar-soporte/<nombre>', methods=['POST'])
def eliminar_soporte(nombre):
    if 'usuario' not in session or session.get('rol') != 'admin':
        return redirect(url_for('main.login'))

    ruta = os.path.join(current_app.config['UPLOAD_FOLDER'], nombre)
    if os.path.exists(ruta):
        os.remove(ruta)
        flash(f"🗑 Se eliminó correctamente el archivo: {nombre}")
    else:
        flash("❌ Archivo no encontrado.")

    return redirect(url_for('main.documentos_subidos'))

@main.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('main.home'))

# =================== FUNCIÓN AUXILIAR ===================

def actualizar_hoja_vida(equipo_id, fecha, descripcion):
    carpeta = os.path.join("static", "hojas_vida")
    archivo_equipo = os.path.join(carpeta, f"{equipo_id}.xlsx")

    if not os.path.exists(archivo_equipo):
        print(f"❌ No se encontró la hoja de vida para el equipo: {equipo_id}")
        return

    wb = load_workbook(archivo_equipo)
    ws = wb.active

    fila = 33
    while True:
        celda = ws[f"A{fila}"]
        if not isinstance(celda, MergedCell) and celda.value is None:
            break
        fila += 1

    ws[f"A{fila}"] = fecha
    ws[f"C{fila}"] = descripcion
    ws[f"M{fila}"] = "SV Romero Romero Miguel Ángel"
    wb.save(archivo_equipo)
